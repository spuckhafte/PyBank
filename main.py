# rows: x- coordinate
# columns: y- coordinate
import openpyxl as xl
from datetime import datetime as dt
import time

wb = xl.load_workbook('Database.xlsx')
sheet = wb['Sheet1']
index = {
    1: "Name: ",
    2: "Contact No.: ",
    3: "Pin: ",
    4: "Balance: Rs."
}
important = []


notification_condition = False
log = "Check Logs         (3)"
log2 = "LOGS: "


class Notification:

    def __init__(self, mes1, mes2, check, key):
        self.mes1 = mes1
        self.mes2 = mes2
        self.check = check
        self.key = key

    def edit_outer_text(self):

        popups = int(sheet.cell(self.key, 6).value) - int(sheet.cell(self.key, 5).value)
        if popups > 0 and self.check is not True:
            self.mes1 = f"Check Logs ({popups} new)  (3)"
            return self.mes1

        else:
            self.mes1 = "Check Logs         (3)"
            return self.mes1

    def edit_inner_text(self):
        popups = int(sheet.cell(self.key, 6).value) - int(sheet.cell(self.key, 5).value)
        if popups > 0 and self.check is not True:
            self.mes2 = f"LOGS (check from- {sheet.cell(self.key, 5).value - 5})"
            return self.mes2

        else:
            self.mes2 = "LOGS: "
            return self.mes2


def start():
    while True:
        logsin = input('Login or Signup: ').lower()
        if 'log' in logsin:
            login()
            break
        elif 'sign' in logsin:
            add_account()
            break
        else:
            print('Wrong input, Try Again!\n')
            continue


def check_name(name):
    for row in range(1, sheet.max_row + 1):
        data_names = sheet.cell(row, 1).value
        if name == data_names:
            return True
        if row == sheet.max_row:
            return False


def date_time():
    now = dt.now()
    datetime = now.strftime("%d/%m/%Y ; %H:%M")
    return datetime


def check_password(pin):
    for row in range(2, sheet.max_row + 1):
        data_pins = sheet.cell(row, 3).value
        if pin == data_pins:
            return True
        if row == sheet.max_row:
            return False


def max_col_row(row):
    j = 1
    while True:
        if sheet.cell(row, j).value is None:
            return j - 1
        else:
            j += 1
            continue


def confirm_response(user_response):
    if 'y' == user_response[0]:
        return True
    if 'n' == in user_response[0]:
        return False


def user_details(key):
    print()
    for col in range(1, 5):
        user_data = sheet.cell(key, col).value
        print(index[col] + str(user_data))
    print()


def login():
    while True:
        print("\nLOGIN\n")
        username = input('Enter name: ').upper()
        password = input('Enter pin: ').upper()

        if check_name(username) and check_password(password):
            important.append(username)
            important.append(password)

            key = int(password.split('-').pop(0))

            sheet.cell(key, 6).value = max_col_row(key)

            user_log = f"Logged in: {date_time()}"
            sheet.cell(key, max_col_row(key) + 1).value = user_log
            wb.save('Database.xlsx')
            break
        else:
            print("Pair of credentials doesn't match our data.\n")
            check = False
            time.sleep(0.8)
            while True:
                res = input('Do you want to make a new account for your self: ').lower()
                if 'y' in res:
                    add_account()
                    check = True
                    break
                elif 'n' in res:
                    break
                else:
                    print('Wrong Input, Try again!\n')
                    continue

            if check:
                break
            else:
                continue


def add_account():
    print('\nSIGNUP\n')
    while True:
        time.sleep(0.8)
        new_name = input('Enter name: ')
        if new_name == "stop":
            print('terminate\n')
            time.sleep(0.8)
            login()
            return

        new_phone = input('Enter Phone Number: ')
        if new_phone == "stop":
            print('terminate\n')
            time.sleep(0.8)
            login()
            return
        try:
            phone_list = []
            for row in range(1, sheet.max_row + 1):
                user_phones = sheet.cell(row, 2).value
                phone_list.append(user_phones)

            if len(new_phone) == 10:
                int(new_phone)
                if int(new_phone) not in phone_list:
                    pass
                else:
                    time.sleep(0.8)
                    print('This phone number is already linked with an account, Try again!\n')
                    continue
            else:
                time.sleep(0.8)
                print("Phone no.'s format is wrong, Try again!\n")
                continue

            new_pin_suffix = input('Enter 4-digit pin: ')
            if new_pin_suffix == "stop":
                print('terminate\n')
                time.sleep(0.8)
                login()
                return

            if len(new_pin_suffix) == 4:
                int(new_pin_suffix)
                new_pin_suffix_again = input('Confirm the new pin: ')
                if new_pin_suffix_again == "stop":
                    print('terminate\n')
                    time.sleep(0.8)
                    login()
                    return
                if new_pin_suffix == new_pin_suffix_again:
                    key = sheet.max_row + 1
                    new_pin = str(key) + '-' + new_pin_suffix
                    time.sleep(0.8)
                    print(f'\nYour new pin is: {new_pin}')
                    time.sleep(0.8)

                    important.append(new_name)
                    important.append(new_pin)

                    sheet.cell(key, 1).value = new_name.upper()
                    sheet.cell(key, 2).value = int(new_phone)
                    sheet.cell(key, 3).value = new_pin
                    sheet.cell(key, 4).value = 0
                    
                    sheet.cell(key, 5).value = sheet.cell(key, 6).value = 0

                    user_log = f"Accounted Created: {date_time()}"
                    sheet.cell(key, max_col_row(key) + 3).value = user_log

                    wb.save('Database.xlsx')
                    print('\nAccount Added!')

                    user_details(key)
                    break
                else:
                    time.sleep(0.8)
                    print("Pins doesn't match, Try Again!\n")
                    continue
            else:
                time.sleep(0.8)
                print("Format of the pin-suffix is wrong (eg. 4444), Try again!\n")
        except:
            time.sleep(0.8)
            print('Format of the input is wrong, Try Again!\n')
            continue


def delete_account(key, pin):
    while True:
        print('\nACCOUNT DELETE PROCESS')
        time.sleep(0.8)
        del_pass = input('Confirm your pin: ')
        if del_pass == "stop":
            print('terminate\n')
            time.sleep(0.8)
            return

        if del_pass == pin:
            warn = input('Are you sure, your account will be permanently deleted: ').lower()
            if 'y' in warn:
                for column in range(1, max_col_row(key) + 1):
                    sheet.cell(key, column).value = None
                wb.save('Database.xlsx')
                time.sleep(0.8)
                add_account()
                break
            elif 'n' in warn:
                break
            else:
                time.sleep(0.8)
                print('Wrong Input, Try Again!')
                continue
        else:
            time.sleep(0.8)
            print("The pin you entered is wrong, Try again!")
            continue


def deposit(key):
    print('\nDEPOSIT MONEY')
    while True:
        password = input('Confirm pin: ')
        if password == "stop":
            print('terminate\n')
            time.sleep(0.8)
            return
        if check_password(password):
            while True:
                try:
                    dep_amount = int(input('Enter the amount to be deposit: Rs.'))
                    if str(dep_amount) == "stop":
                        print('terminate!\n')
                        time.sleep(0.8)
                        return
                except:
                    time.sleep(0.8)
                    print('Format of the amount is wrong, Try again!\n')
                    continue

                confirm = input(f'\nAre you sure, Rs.{str(dep_amount)} will be deposited to your account: ')
                if confirm_response(confirm):
                    sheet.cell(key, 4).value += dep_amount
                    time.sleep(0.8)
                    print('Money deposited, check Logs!\n')

                    user_balance = sheet.cell(key, 4).value
                    user_log = f"Amount Deposited: Rs.{dep_amount}; Your Balance: Rs.{user_balance}; On: {date_time()}"
                    sheet.cell(key, max_col_row(key) + 1).value = user_log
                    wb.save('Database.xlsx')
                    return

                elif not confirm_response(confirm):
                    time.sleep(0.8)
                    print('Deposit Cancelled\n')

                    user_log = f"Deposit Failed: {date_time()}"
                    sheet.cell(key, max_col_row(key) + 1).value = user_log
                    wb.save('Database.xlsx')
                    return
                else:
                    time.sleep(0.8)
                    print('\nWrong input, Try again!')
        else:
            time.sleep(0.8)
            print('Wrong pin, Try again!')
            continue


def transfer_money(key):
    print('\nTRANSFER MONEY')
    while True:
        try:
            time.sleep(0.8)
            target_phone = input('Enter the phone number of the user, to whom you want to transfer money: ')
            if target_phone == "stop":
                print('terminate!\n')
                time.sleep(0.8)
                return
            int(target_phone)
        except:
            time.sleep(0.8)
            print('Format for the phone number is wrong, Try again!\n')
            continue

        if len(target_phone) > 10 or len(target_phone) < 10:
            time.sleep(0.8)
            print('Format for the phone number is wrong, Try again!\n')
            continue

        target_row = 1
        condition = False
        for cell in range(2, sheet.max_row + 1):
            target_row += 1
            phone = sheet.cell(cell, 2).value
            if int(phone) == int(target_phone):
                condition = True
                break
            else:
                continue

        if not condition:
            time.sleep(0.8)
            print("The phone number doesn't match to any account, Try again!\n")
            continue

        if target_phone == sheet.cell(key, 2).value:
            time.sleep(0.8)
            print("You can't transfer money to yourself, Try again!\n")
            continue

        while True:
            time.sleep(0.8)
            pin = input('Confirm your pin: ')
            if pin == "stop":
                print('terminate!\n')
                time.sleep(0.8)
                return

            if check_password(pin):
                try:
                    transfer_amount = int(input('Enter the amount to be transferred: Rs.'))
                    if transfer_amount == "stop":
                        print('terminate!\n')
                        time.sleep(0.8)
                        return
                    if transfer_amount > sheet.cell(key, 4).value or transfer_amount <= 0:
                        time.sleep(0.8)
                        print('Invalid amount, Try again!\n')
                        continue
                except:
                    time.sleep(0.8)
                    print('Kindly enter the amount correctly, Try again!\n')
                    continue

                while True:
                    time.sleep(0.8)
                    confirm = input(
                        f'\nAre you sure, Rs.{str(transfer_amount)} will be transferred to an account linked '
                        f'to the phone number {target_phone}: ')

                    if confirm_response(confirm):
                        sheet.cell(target_row, 4).value += transfer_amount
                        sheet.cell(key, 4).value -= transfer_amount

                        user_balance = sheet.cell(key, 4).value
                        target_balance = sheet.cell(target_row, 4).value

                        time.sleep(0.8)
                        print('\nAmount Transferred!')
                        print('Check your Logs for details.\n')

                        now = dt.now()
                        datetime = now.strftime("%d/%m/%Y ; %H:%M")
                        username = sheet.cell(key, 1).value
                        userphone = sheet.cell(key, 2).value

                        user_log = f"Amount transferred: Rs.{transfer_amount}; Target phone no.: {target_phone}; Your Balance: Rs.{user_balance}; On: {datetime}"
                        target_log = f"Amount received: Rs.{transfer_amount}; From: {username}({userphone}); Your Balance: Rs.{target_balance}; On: {datetime}"

                        sheet.cell(key, max_col_row(key) + 1).value = user_log
                        sheet.cell(target_row, max_col_row(target_row) + 1).value = target_log

                        wb.save('Database.xlsx')
                        return

                    elif not confirm_response(confirm):
                        time.sleep(0.8)
                        print('Transaction Failed!\n')

                        user_log = f"Transaction Failed: {date_time()}"
                        sheet.cell(key, max_col_row(key) + 1).value = user_log
                        wb.save('Database.xlsx')
                        return
                    else:
                        time.sleep(0.8)
                        print('Wrong input, Try again!\n')
                        continue


def change_password(pin, key):
    print('CHANGE PASSWORD')
    print("\nNote: You can't change the prefix\n")
    while True:
        time.sleep(0.8)
        old_pass = input('Enter old pin (with prefix): ')
        if old_pass == "stop":
            print('terminate!\n')
            time.sleep(0.8)
            return
        if old_pass == pin:
            break
        else:
            print('Wrong pin, Try again!\n')
            continue

    while True:
        time.sleep(0.8)
        new_pass = input('Enter new pin (without prefix): ')
        if new_pass == "stop":
            print('terminate!\n')
            time.sleep(0.8)
            return
        if type(int(new_pass)) == int and len(new_pass) == 4:
            new_pass_again = input('Confirm the new pin: ')
            if new_pass_again == "stop":
                print('terminate!\n')
                time.sleep(0.8)
                return
            if new_pass == new_pass_again:
                sheet.cell(key, 3).value = str(key) + '-' + new_pass

                user_log = f"Pin changed: from- {old_pass}, to- {sheet.cell(key, 3).value}; On: {date_time()}"
                sheet.cell(key, max_col_row(key) + 1).value = user_log
                wb.save('Database.xlsx')

                time.sleep(0.8)
                print(f'\nNew Pin: {sheet.cell(key, 3).value}\n')
                break
            else:
                time.sleep(0.8)
                print("Pins doesn't match, Try Again!\n")
                continue
        else:
            time.sleep(0.8)
            print("Format of the pin-suffix is wrong (eg. 4444), Try again!\n")
            continue


# Start of interaction

start()
clientname = important[0]
userpin = important[1]
keyword = int(userpin.split('-').pop(0))

print(f"\nWelcome to the PyBank, {clientname.upper()}\n")
while True:
    log2 = Notification(mes1=log, mes2=log2, check=notification_condition, key=keyword).edit_outer_text()
    time.sleep(0.4)
    print(f'''Contents:
    Check Details      (1)
    Transfer Money     (2)
    {log2}
    Deposit Money      (4)
    Change Pin         (pin)
    Delete Account     (del)
    Logout             (out)
    Terminate Process  (stop)
''')

    user_input = input('> ')
    if user_input == "1":
        user_details(keyword)

    elif user_input == "2":
        transfer_money(keyword)

    elif user_input == "3":
        log = Notification(mes1=log, mes2=log2, check=notification_condition, key=keyword).edit_inner_text()
        print(f'\n{log}')
        if notification_condition is not True:
            time.sleep(1.5)
        else:
            time.sleep(0.8)
        for i in range(7, max_col_row(keyword) + 1):
            print(f'{i - 6}- {sheet.cell(keyword, i).value}')
        notification_condition = True

        cont = input('Enter any key to continue: ')
        print()
        time.sleep(0.2)

    elif user_input == "4":
        deposit(keyword)

    elif user_input == "pin":
        change_password(userpin, keyword)

    elif user_input == "del":
        delete_account(keyword, userpin)
        continue

    elif user_input == "out":
        time.sleep(0.5)
        print('\nThank you for visiting us!')

        sheet.cell(keyword, 5).value = max_col_row(keyword)
        wb.save('Database.xlsx')
        notification_condition = False

        break

    elif user_input == "stop":
        time.sleep(0.5)

    else:
        print('Wrong input, Try again!\n')
